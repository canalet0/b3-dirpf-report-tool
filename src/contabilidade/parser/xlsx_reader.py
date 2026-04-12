from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _is_total_row(row: list[str | None]) -> bool:
    non_empty = [c for c in row if c is not None and str(c).strip() != ""]
    if len(non_empty) <= 2:
        for cell in non_empty:
            if str(cell).strip().lower() == "total":
                return True
    return False


def _is_empty_row(row: list[str | None]) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def _sheet_to_rows(ws: Worksheet) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row in ws.iter_rows(values_only=True):
        cells: list[str | None] = [
            str(c).strip() if c is not None else None for c in row
        ]
        rows.append(cells)

    # Strip trailing empty rows
    while rows and _is_empty_row(rows[-1]):
        rows.pop()

    # Strip trailing total rows
    while rows and _is_total_row(rows[-1]):
        rows.pop()

    return rows


def read_xlsx(path: Path) -> dict[str, list[list[str | None]]]:
    wb: Workbook = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, list[list[str | None]]] = {}
    for name in wb.sheetnames:
        ws: Worksheet = wb[name]
        result[name] = _sheet_to_rows(ws)
    return result
